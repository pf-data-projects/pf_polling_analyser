""" 
This file handles the logic for building and styling the excel table
outputs for the whole program. It takes trimmed table data as an
input. The logic broadly follows these steps.

1. call external functions to generate a cover page and contents page.
2. create a writer object using pandas/xlsxwriter to create the main results.
3. Add contents and cover pages to excel sheet.
4. Create individual sheets for all questions in the table.
5. Create links to each table and from each table back to contents in
each sheet.
6. cache the excel file for download later after the function has returned.

#### AS OF 01/12/2023 ####
7. re-opens cached file and carries out following edits using openpyxl.
    - add pf logo to top corners of all sheets.
    - add crossbreak headers to crossbreak groups.
    - Add some custom styles to these new rows/cols.
"""

import io
import math

import pandas as pd
import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font, Alignment, Border
from django.core.cache import cache

from .contents import create_contents_page
from .cover import create_cover_page
from .helper import get_column_letter

def create_workbook(
    request, data, questions_list, grids, unique_ids,
    title, dates, comments, start, end, id_column, rebased_headers):
    """
    The function that controls the creation and formatting of
    polling tables.
    """
    # create a further trimmed dataframe for excel output
    trimmed_data = data.drop(
        ['IDs', 'Types', 'Base Type', 'Rebase comment needed'],
        axis=1
    )

    trimmed_data.iat[0, 0] = "Unweighted"

    # ensure blank cols remain blank
    for col in trimmed_data.columns:
        if trimmed_data[col].iloc[0] == 0:
            trimmed_data[col].iloc[0] = ""
            trimmed_data[col].iloc[1] = ""
            trimmed_data = trimmed_data.rename(columns={col: ''})

    # create cover page and contents page
    # blank = {'Table of contents'}
    cover_df = create_cover_page(data, title, dates)
    contents_df = create_contents_page(data, questions_list, comments, grids, unique_ids, id_column)

    # define variables for caching
    cache_key = "tables_for_user_" + str(request.user.id)
    output = io.BytesIO()

    # Format the tables with xlsxwriter and pandas
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ create main results polling table
        cover_df.to_excel(writer, index=False, sheet_name="Cover Page")
        contents_df[0].to_excel(writer, index=False, sheet_name="Contents")
        trimmed_data.to_excel(writer, index=False, sheet_name='Full Results')
        workbook = writer.book
        totals_format = workbook.add_format({
            'num_format': '0',
            'align': 'center',
        })
        weighted_totals_format = workbook.add_format({
            'num_format': '0',
            'align': 'center',
        })
        weighted_totals_format.set_bottom(1)
        questions_border = workbook.add_format({
            "align": "left"
        })
        questions_border.set_right(1)
        weight_label_format = workbook.add_format({
            "align": "right"
        })

        # define results sheet and add basic styles
        results_sheet = writer.sheets['Full Results']
        results_sheet.hide_gridlines(2)
        results_sheet.freeze_panes(5, 1)

        # define contents sheet and add basic styles
        contents_sheet = writer.sheets['Contents']
        contents_sheet.hide_gridlines(2)
        contents_sheet.set_column(2, 2, 100)
        centre_format = workbook.add_format({
            'align': 'center',
        })
        contents_sheet.set_column("D:D", 20, centre_format)
        # contents_sheet.set_column(3, 3, 30)
        contents_sheet.set_column(4, 4, 40)
        right_format = workbook.add_format({
            'align': 'right',
        })
        contents_sheet.set_column("B:B", None, right_format)

        # define cover sheet and add basic styles
        cover_sheet = writer.sheets['Cover Page']
        cover_sheet.hide_gridlines(2)
        cover_sheet.set_column(3, 3, 60)
        cover_sheet.set_zoom(115)

        # Create links to full results table rows in contents page
        updated_list = questions_list[1:]
        prev_question = 0
        for question in updated_list:
            row_index = contents_df[0].index[contents_df[0]['Row in Full Results'] == question]
            row_index.tolist()
            excel_col = 'D'
            excel_row = row_index[0] + 2
            cell = f"{excel_col}{excel_row}"
            contents_sheet.write_url(
                cell,
                f"internal:'Full Results'!A{prev_question}",
                string=f"row {question}"
            )
            prev_question = question

        results_sheet.set_zoom(85)
        header_format = workbook.add_format({
            "bg_color": "#FFA500",
            "bold": True,
            "font_color": "#FFFFFF",
            "align": "center"
        })
        percent_format = workbook.add_format({
            'num_format': '0%',
            'align': 'center'
        })
        question_format = workbook.add_format({"bold": True})
        # Apply a general format to the entire column
        # without the percentage format
        results_sheet.set_column(1, len(data.columns) - 1, 15)
        results_sheet.set_column(0, 0, 80, cell_format=questions_border)

        # round weighted totals to nearest integer
        row_as_list = trimmed_data.iloc[1].values.tolist()
        for col, number in enumerate(row_as_list, start=0):
            if isinstance(number, str):
                results_sheet.write(2, col, number)
            else:
                results_sheet.write_number(2, col, number, weighted_totals_format)

        row_as_list_totals = trimmed_data.iloc[0].values.tolist()
        for col, number in enumerate(row_as_list_totals, start=0):
            if isinstance(number, str):
                results_sheet.write(1, col, number)
            else:
                results_sheet.write_number(1, col, number, totals_format)

        # create question style and loop to apply them
        for i in range(2, len(data)):
            question_value = data.iloc[i, 3]
            if data.at[i, 'Base Type'] == 'Question' or data.at[i, 'Base Type'] == 'sub_Question':
                results_sheet.write(i + 1, 0, question_value, question_format)

        # apply percentage format to data.
        format_percentages(trimmed_data, results_sheet, percent_format)

        for col_num, value in enumerate(trimmed_data.columns.values):
            results_sheet.write(0, col_num, value, header_format)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ create individual tables.
        # ~~~~~~ Apologies for how cursed this bit of code is atm.
        non_header_data = data.iloc[2:]
        header = data.loc[0:1]
        ids = non_header_data['IDs'].tolist()
        checked = []
        for qid in ids:
            # add a grid summary to the table if it exists
            if qid in unique_ids:
                idx = unique_ids.index(qid)
                grid_summary = grids[idx]
                grid_summary.to_excel(
                    writer,
                    index=False,
                    sheet_name=f'Grid Summary - {qid}'
                )
            if qid != "" and qid not in checked:
                # create table containing just 1 question
                sub_table = data[(data['IDs'] == qid)]
                # add a space beneath where header goes
                rows_list = []
                for index, row in sub_table.iterrows():
                    if row['Base Type'] in ['Question']:
                        empty_row = pd.Series(
                            [''] * len(data.columns), index=data.columns)
                        rows_list.append(empty_row)
                    rows_list.append(row)
                sub_table = pd.concat(
                    [pd.DataFrame([row]) for row in rows_list],
                    ignore_index=True
                )
                # concatinate the headers back to the top of the sub table
                concat_sub_table = pd.concat(
                    [header, sub_table],
                    ignore_index=True
                )
                sub_table.reset_index(drop=True, inplace=True)
                # drop columns that we don't want to see in excel
                concat_sub_table = concat_sub_table.drop(
                    ['IDs', 'Types', 'Base Type', 'Rebase comment needed'],
                    axis=1
                )
                # # add a new row for link back to contents page
                # new_row = {concat_sub_table.columns[0]: 'Back to contents'}
                # for col in concat_sub_table.columns[1:]:
                #     new_row[col] = ''
                # concat_sub_table.loc[len(concat_sub_table)] = new_row

                # ensure blank cols remain blank
                for col in concat_sub_table.columns:
                    if concat_sub_table[col].iloc[0] == 0:
                        concat_sub_table[col].iloc[0] = ""
                        concat_sub_table[col].iloc[1] = ""
                        concat_sub_table = concat_sub_table.rename(columns={col: ''})

                # Add the table to excel with sheetname based on QID
                concat_sub_table.to_excel(
                    writer,
                    index=False,
                    sheet_name=f'Question ID - {qid}'
                )
                # define sheet for formatting
                question_sheet = writer.sheets[f'Question ID - {qid}']
                question_sheet.set_column(1, len(data.columns) - 1, 15)
                question_sheet.set_column(0, 0, 80, cell_format=questions_border)
                question_sheet.hide_gridlines(2)
                question_sheet.freeze_panes(5, 1)
                # format numbers to nice percentages
                format_percentages(
                    concat_sub_table, question_sheet, percent_format
                )

                # Do the same for grid summaries
                if qid in unique_ids:
                    grid_sheet = writer.sheets[f'Grid Summary - {qid}']
                    grid_sheet.set_column(1, len(grid_summary.columns) - 1, 40)
                    grid_sheet.set_column(0, 0, 80, cell_format=questions_border)
                    grid_sheet.hide_gridlines(2)
                    grid_sheet.freeze_panes(2, 1)
                    format_percentages(
                        grid_summary, grid_sheet, percent_format
                    )

                # format questions
                for i in range(len(sub_table)):
                    question_value = concat_sub_table.iloc[i + 2, 0]
                    if sub_table.loc[i, 'Base Type'] == 'Question' or sub_table.loc[i, 'Base Type'] == 'sub_Question':
                        question_sheet.write(
                            i + 3,
                            0,
                            question_value,
                            question_format
                        )

                # format headers
                for col, value in enumerate(concat_sub_table.columns.values):
                    question_sheet.write(0, col, value, header_format)

                for col, number in enumerate(row_as_list, start=0):
                    if isinstance(number, str):
                        question_sheet.write(2, col, number)
                    else:
                        question_sheet.write_number(2, col, number, weighted_totals_format)
                for col, number in enumerate(row_as_list_totals, start=0):
                    if isinstance(number, str):
                        question_sheet.write(1, col, number)
                    else:
                        question_sheet.write_number(1, col, number, totals_format)
                checked.append(qid)

    output.seek(0)
    cache.set(cache_key, output.getvalue(), timeout=300)


    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ EDITS WITH OPENPYXL
    cached_file_content = cache.get(cache_key)

    # Create a BytesIO object from your cached content
    file_obj = io.BytesIO(cached_file_content)

    # Load the workbook from the file object and define image path
    wb = load_workbook(file_obj)
    img_path = 'static/assets/pf.jpg'

    # Add pf logo to top of all sheets
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.insert_rows(1)
        ws.row_dimensions[1].height = 60

        img = Image(img_path)
        scale_x = 0.5
        scale_y = 0.5
        img.width = img.width * scale_x
        img.height = img.height * scale_y
        if sheet == 'Cover Page':
            ws.add_image(img, 'D1')
        else:
            ws.add_image(img, 'A1')

    # Add extra space in cover page.
    cover_page = wb['Cover Page']
    for i in range(1, 4):
        cover_page.insert_cols(idx=1)
    cover_page['D2'].border = Border(
        left=None, right=None, top=None, bottom=None)
    cover_page['D10'].alignment = Alignment(wrapText=True)
    cover_page['D12'].alignment = Alignment(wrapText=True)

    # Add extra row and headers for the standard crossbreaks.
    protected_sheets = [
        'Cover Page',
        'Contents',
    ]

    # create links to all the tables in the contents page.
    contents_sheet = wb['Contents']
    for cell in contents_sheet['C']:
        if id_column:
            qid = contents_sheet[f'B{cell.row}']
            if cell.value is None:
                continue
            if 'Grid' and 'Question' not in cell.value:
                # print(f"#'Question ID - {qid.value}'!A1")
                cell.hyperlink = f"#'Question ID - {qid.value}'!A1"
            if 'Grid' in cell.value:
                cell.hyperlink = f"#'{cell.value}'!A1"
        else:
            qid = contents_sheet[f'A{cell.row}']
            if cell.value is None:
                continue
            if 'Grid' and 'Question' not in cell.value:
                # print(f"#'Question ID - {qid.value}'!A1")
                cell.hyperlink = f"#'Question ID - {qid.value}'!A1"
            if 'Grid' in cell.value:
                cell.hyperlink = f"#'{cell.value}'!A1"

    # remove any n/a values from the contents Full Results column
    # and remove id column if user has not chosen it.
    ws = wb['Contents']
    for cell in ws['D']:
        if cell.value == 'n/a':
            cell.value = None
    if not id_column:
        for cell in ws['A']:
            cell.value = None
    img = Image(img_path)
    scale_x = 0.5
    scale_y = 0.5
    img.width = img.width * scale_x
    img.height = img.height * scale_y
    ws.add_image(img, 'A1')

    for sheet in wb.sheetnames:
        if 'Grid' in sheet:
            protected_sheets.append(sheet)

    for sheet in wb.sheetnames:
        if sheet not in protected_sheets:
            ws = wb[sheet]
            max_column = ws.max_column
            for row in [3, 4]:
                for col in range(1, max_column + 1):
                    c = ws.cell(row=row, column=col)
                    try:
                        if c.value == 0 or c.value == "0":
                            c.value = None
                    except AttributeError as e:
                        print(e)

    # add standard cb headers.
    for sheet in wb.sheetnames:
        if sheet not in protected_sheets:
            ws = wb[sheet]
            ws.insert_rows(2)
            ws.row_dimensions[2].height = 40
            cols = trimmed_data.columns
            if "Male" in cols:
                excel_coord = get_header_coords("Male", trimmed_data)
                excel_coord2 = get_header_coords("Female", trimmed_data)
                ws[excel_coord] = "Gender"
                ws.merge_cells(f"{excel_coord}:{excel_coord2}")
            if "18-24" in cols:
                excel_coord = get_header_coords("18-24", trimmed_data)
                excel_coord2 = get_header_coords("65+", trimmed_data)
                ws[excel_coord] = "Age"
                ws.merge_cells(f"{excel_coord}:{excel_coord2}")
            if "London" in cols:
                excel_coord = get_header_coords("London", trimmed_data)
                excel_coord2 = get_header_coords("Northern Ireland", trimmed_data)
                ws[excel_coord] = "Region"
                ws.merge_cells(f"{excel_coord}:{excel_coord2}")
            if "AB" in cols:
                excel_coord = get_header_coords("AB", trimmed_data)
                excel_coord2 = get_header_coords("DE", trimmed_data)
                ws[excel_coord] = "Socio-Economic Grade"
                ws.merge_cells(f"{excel_coord}:{excel_coord2}")
            if "Yes" in cols:
                excel_coord = get_header_coords("Yes", trimmed_data)
                excel_coord2 = get_header_coords("No", trimmed_data)
                ws[excel_coord] = "Has Children?"
                ws.merge_cells(f"{excel_coord}:{excel_coord2}")
            if "No children" in cols:
                excel_coord = get_header_coords("No children", trimmed_data)
                excel_coord2 = get_header_coords("Yes - child/children over 18 years old", trimmed_data)
                ws[excel_coord] = "Has Children?"
                ws.merge_cells(f"{excel_coord}:{excel_coord2}")
            if "GCSE or equivalent (Scottish National/O Level)" in cols:
                excel_coord = get_header_coords(
                    "GCSE or equivalent (Scottish National/O Level)", 
                    trimmed_data
                )
                excel_coord2 = get_header_coords("Doctorate (PhD/DPHil)", trimmed_data)
                ws[excel_coord] = "Highest Level of Education"
                ws.merge_cells(f"{excel_coord}:{excel_coord2}")
            if "The Brexit Party" in cols:
                excel_coord = get_header_coords("The Brexit Party", trimmed_data)
                excel_coord2 = get_header_coords("Liberal Democrat", trimmed_data)
                ws[excel_coord] = "2019"
                ws.merge_cells(f"{excel_coord}:{excel_coord2}")
            if "brexit_Leave" in cols:
                excel_coord = get_header_coords("brexit_Leave", trimmed_data)
                excel_coord2 = get_header_coords("brexit_I did not vote", trimmed_data)
                ws[excel_coord] = "EU 2016 Vote"
                ws.merge_cells(f"{excel_coord}:{excel_coord2}")
            if "intention_Conservative" in cols:
                excel_coord = get_header_coords("intention_Conservative", trimmed_data)
                excel_coord2 = get_header_coords("intention_Liberal Democrats", trimmed_data)
                ws[excel_coord] = "Voting Intention"
                ws.merge_cells(f"{excel_coord}:{excel_coord2}")


    # Add headers for non-standard crossbreaks

    non_standard = [col for col in trimmed_data.columns if ":" in col]

    for sheet in wb.sheetnames:
        if sheet not in protected_sheets:
            ws = wb[sheet]
            cols_done = []
            for col in non_standard:
                header_coords = get_header_coords(col, trimmed_data)
                title_coords = get_title_coords(col, trimmed_data)
                header = col.split(":")[0]
                col_title = col.split(":")[1]
                cols_count = len([col for col in non_standard if header in col])
                if header not in cols_done:
                    ws[header_coords] = header
                    ws[title_coords] = col_title
                    last_col_coords = shift_right(header_coords, cols_count - 1)
                    ws.merge_cells(f'{header_coords}:{last_col_coords}')
                else:
                    ws[title_coords] = col_title
                cols_done.append(header)

    # Add styles to the crossbreak headers.
    # (and grid headers)

    fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    font = Font(bold=True, color='FFFFFF', size=14)
    smaller_font = Font(bold=True, color='FFFFFF', size=11)
    alignment = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left')

    for sheet in wb.sheetnames:
        if sheet not in protected_sheets:
            ws = wb[sheet]
            ws['A3'] = " "
            for cell in ws['2']:
                cell.fill = fill
                cell.font = font
                cell.alignment = alignment
            # remove brexit and intention prefixes in final table
            for cell in ws['3']:
                if 'brexit_' in str(cell.value):
                    cleaned_value = str(cell.value).replace('brexit_', "")
                    cell.value = cleaned_value
                if 'intention_' in str(cell.value):
                    cleaned_value = str(cell.value).replace('intention_', "")
                    cell.value = cleaned_value
        if 'Grid' in sheet:
            ws = wb[sheet]
            for cell in ws['2']:
                cell.fill = fill
                cell.font = smaller_font
                cell.alignment = alignment
            ws['A2'].alignment = left_align

    # format grid numbers as percentages.
    for qid in unique_ids:
        try:
            grid = wb[f'Grid Summary - {qid}']
            idx = unique_ids.index(qid)
            grid_summary = grids[idx]
            for row in range(3, len(grid_summary) + 3):
                for col in range(2, len(grid_summary.columns) + 2):
                    cell = grid.cell(row=row, column=col)
                    if cell.value is not None:
                        cell.value = round(cell.value)
                        cell.number_format = '0%'
                        cell.value = cell.value / 100
        except KeyError:
            print("This grid does not exist")


    # Add the mini titles to the full results sheet, and then to all sheets.

    title_font = Font(bold=True, size=25)
    title_font_smaller = Font(bold=True, size=18)
    title_align = Alignment(vertical='center')

    full_results = wb['Full Results']
    full_results['C1'] = 'Full Results'
    full_results['C1'].font = title_font
    full_results['C1'].alignment = title_align

    contents_page = wb['Contents']
    contents_page['C3'].hyperlink = "#'Full Results'!B2"
    contents_page['C3'].value = 'Full Results Table'
    contents_page['E3'].value = ''

    protected_sheets.append('Full Results')

    for sheet in wb.sheetnames:
        if sheet not in protected_sheets:
            ws = wb[sheet]
            ws['B1'] = ws['A7'].value
            ws['B1'].font = title_font_smaller
            ws['B1'].alignment = title_align
            ws['A3'].hyperlink = '#Contents!A1'
            ws['A3'].value = 'Back to Contents'

    wb = update_column_headers(wb, rebased_headers)

    output = io.BytesIO()
    wb.save(output)

    cache.set(cache_key, output.getvalue(), timeout=300)
    return cache_key

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ EXTRA HELPER FUNCTIONS HERE

def format_percentages(data, sheet, cell_format):
    """
    Loop through rows, starting from the fourth row, 
    and apply the percentage format.
    """
    for row_num in range(3, len(data)):  # start from the fourth row (index 3)
        row_data = data.iloc[row_num]
        for col_num in range(1, len(data.columns)):
            col_name = data.columns[col_num]
            if col_name != " ":
                cell_value = data.iloc[row_num, col_num]
                # Check if the cell contains a number (int or float)
                if isinstance(cell_value, (int, float)):
                    if not math.isnan(cell_value):
                        # Apply percent format to the cell because it contains a number
                        sheet.write_number(
                            row_num + 1, col_num, cell_value, cell_format)
                elif pd.isna(cell_value) or cell_value == '':
                    # If the cell is NaN or an empty string, write an empty string
                    sheet.write_string(row_num + 1, col_num, '')
                else:
                    # Otherwise, write the value as it is
                    sheet.write(row_num + 1, col_num, cell_value)

def get_header_coords(colname, data):
    """
    Gets the excel coordinates for any
    given crossbreak headers
    """
    col_index = data.columns.get_loc(colname)
    excel_col = get_column_letter(col_index + 1)
    excel_coord = excel_col + '2'
    return excel_coord

def get_title_coords(colname, data):
    """
    Gets the excel coordinates for any
    given crossbreak headers
    """
    col_index = data.columns.get_loc(colname)
    excel_col = get_column_letter(col_index + 1)
    excel_coord = excel_col + '3'
    return excel_coord

def shift_right(cell, x):
    """
    Gets the excel coordinates
    x to the right of the coordinates entered.
    """
    # Function to convert a column letter to a number (e.g., 'A' -> 1, 'B' -> 2)
    def col_to_num(col_str):
        num = 0
        for c in col_str:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num

    # Function to convert a number back to a column letter (e.g., 1 -> 'A', 2 -> 'B')
    def num_to_col(n):
        col_str = ''
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            col_str = chr(65 + remainder) + col_str
        return col_str

    # Split the cell address into column and row parts
    col = ''.join(filter(str.isalpha, cell))
    row = ''.join(filter(str.isdigit, cell))

    # Shift the column
    new_col_num = col_to_num(col) + x
    new_col = num_to_col(new_col_num)
    return new_col + row

def num_to_col(n):
    """
    helper func to convert
    an int to excel col letter.
    """
    col_str = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        col_str = chr(65 + remainder) + col_str
    return col_str

def update_column_headers(workbook, rebased_headers):
    """
    Loops through the rebase headers dictionary and applies the data
    to the relevant sheet.
    """
    # iterate through question ids
    for key, value in rebased_headers.items():
        col = 2
        # skip questions that have been trimmed off
        try:
            sheet = workbook[f'Question ID - {key}']
        except KeyError:
            print("This question does not exist in the trimmed data.")
            continue
        # Iterate through crossbreaks
        for k, v in value.items():
            column = num_to_col(col)
            # print(column, type(sheet[f'{column}3'].value))
            if sheet[f'{column}3'].value == '' or sheet[f'{column}3'].value is None:
                col += 1
                column = num_to_col(col)
            sheet[f'{column}3'].value = k
            sheet[f'{column}4'].value = v[0]
            sheet[f'{column}5'].value = v[1]
            col += 1
    return workbook
